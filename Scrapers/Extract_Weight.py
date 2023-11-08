import openpyxl
import re

# Load the Excel workbook
file_path = "C:\Desktop\WebScrapers\Outputs\Truefarm.xlsx"
workbook = openpyxl.load_workbook("C:\Desktop\WebScrapers\Outputs\Truefarm.xlsx")
worksheet = workbook.active

# Define a regular expression pattern to extract weight
pattern = re.compile(r'(\d+\s?[gkGK][gG]?)')

# Iterate over the strings in the Excel file
for row in range(2,20):  # Assuming data starts from row 2 and you have 310 strings
    cell_value = worksheet.cell(row=row, column=2).value  # Assuming strings are in column 1
    match = pattern.search(cell_value)
    if match:
        weight = match.group(1)
        # Store the extracted weight in column 2 (or any other column you prefer)
        worksheet.cell(row=row, column=4).value = weight

# Save the updated workbook
workbook.save("updated_excel_file3.xlsx")   #Put the name of the Excel doc you want to create
