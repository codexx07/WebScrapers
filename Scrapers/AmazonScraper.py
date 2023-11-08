import requests
from bs4 import BeautifulSoup
import openpyxl

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
}

base_url = "https://www.amazon.in/s?k=Dry+fruits+and+nuts&i=everyday-essentials&srs=5756143031&bbn=5756143031&crid=GZU02H467OAI&qid=1698412754&sprefix=dry+fruits+and+n%2Ceveryday-essentials%2C470&ref=sr_pg_1"

# Create an Excel workbook and add a worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet["A1"] = "Product Name"
worksheet["B1"] = "Product Price"

row_counter = 2

for page_num in range(0, 20):  # Adjust the range as needed
    response = requests.get(base_url, headers=headers)
    soup = BeautifulSoup(response.content, "html.parser")
    divsname = soup.find_all("div", class_="a-section a-spacing-none puis-padding-right-small s-title-instructions-style")
    divs = soup.find_all("div", class_="a-section a-spacing-none a-spacing-top-micro puis-price-instructions-style")

    for div_name, div_price in zip(divsname, divs):
        # Extract product names
        product_name = div_name.find("span", class_="a-size-medium a-color-base a-text-normal")
        if product_name:
            worksheet.cell(row=row_counter, column=1, value=product_name.text.strip())
        else:
            worksheet.cell(row=row_counter, column=1, value="N/A")

        # Extract product prices
        price_div = div_price.find("span", class_="a-price")
        if price_div:
            product_price = price_div.find("span", class_="a-price-whole")
            if product_price:
                worksheet.cell(row=row_counter, column=2, value=product_price.text.strip())
            else:
                worksheet.cell(row=row_counter, column=2, value="N/A")
        else:
            worksheet.cell(row=row_counter, column=2, value="N/A")
        
        row_counter += 1

    np = soup.find('a', class_="s-pagination-item s-pagination-next s-pagination-button s-pagination-separator")
    if np is not None:
        cnp = "https://www.amazon.in/" + np.get("href")
        base_url = cnp
    else:
        break

workbook.save("Amazon_ProductsList_20_Pages.xlsx")
print("Data has been written to Amazon_ProductsList6.xlsx")
