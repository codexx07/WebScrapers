import requests
from bs4 import BeautifulSoup
import openpyxl

headers = {
    "User-Agent": "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:91.0) Gecko/20100101 Firefox/91.0"
}

base_url = "https://www.flipkart.com/search?q=dry+fruits+and+nuts&otracker=search&otracker1=search&marketplace=FLIPKART&as-show=off&as=off&page=1"

# Create an Excel workbook and add a worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet["A1"] = "Product Name"
worksheet["B1"] = "Product Price"

row_counter = 2

response = requests.get(base_url, headers=headers)
print(response)
soup = BeautifulSoup(response.content, "html.parser")


# for page_num in range(0, 20):  # Adjust the range as needed
#     response = requests.get(base_url, headers=headers)
#     soup = BeautifulSoup(response.content, "html.parser")
#     divs = soup.find_all("div", class_="_4ddWXP")

#     for div in divs:
#         # Extract product names
#         product_name = div.find("a", class_="IRpwTa")
#         # Extract product prices
#         product_price = div.find("div", class_="_30jeq3")

#         if product_name:
#             worksheet.cell(row=row_counter, column=1, value=product_name.text.strip())
#         else:
#             worksheet.cell(row=row_counter, column=1, value="N/A")

#         if product_price:
#             worksheet.cell(row=row_counter, column=2, value=product_price.text.strip())
#         else:
#             worksheet.cell(row=row_counter, column=2, value="N/A")

#         row_counter += 1

#     np = soup.find('a', class_="_1LKTO3")
#     if np is not None:
#         cnp = "https://www.flipkart.com" + np.get("href")
#         base_url = cnp
#     else:
#         break

# workbook.save("Flipkart_ProductsList1.xlsx")
# print("Data has been written to Flipkart_ProductsList_20_Pages.xlsx")
